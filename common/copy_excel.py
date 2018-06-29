from openpyxl import load_workbook
import openpyxl
import os
import xlrd
import xlwt
from xlutils.copy import copy

def copy_excel(excelpath1,excelpath2):
    '''将excel1的数据复制到excel2中'''
    wb2 = openpyxl.Workbook()
    wb2.save(excelpath2)
    #读取数据
    wb1 = openpyxl.load_workbook(excelpath1)
    wb2 = openpyxl.load_workbook(excelpath2)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]
    max_row = sheet1.max_row
    max_col = sheet1.max_column

    for m in list(range(1,max_row+1)):
        for n in list(range(97,97+max_col)):
            n = chr(n)
            i ='%s%d'% (n, m)
            cell1 = sheet1[i].value
            sheet2[i].value = cell1

    wb2.save(excelpath2)
    wb1.close()
    wb2.close()

class Write_excel(object):
    '''修改excel数据'''
    def __init__(self,filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def write(self,coord,value):
        '''写入数据，如(A4，"hello"),第一列第四行写入数据"hello"'''

        self.ws.cell(coord).value = value
        self.wb.save(self.filename)



if __name__ == "__main__":
    current_path = os.path.dirname(os.path.realpath(__file__))
    testdate_path = os.path.join(os.path.dirname(current_path),"testdate")
    report_path = os.path.join(os.path.dirname(current_path),"report")
    #print(current_path)
    #print(testdate_path)
    #print(report_path)
    filePath1 = os.path.join(testdate_path,"test.xlsx")
    filePath2 = os.path.join(report_path,"test_report.xlsx")
    copy_excel(filePath1,filePath2)

    wt = Write_excel(filePath2)
    wt.write("A4","hello")


